import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def load_normalized_data(csv_path):
    rows = []
    with open(csv_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("###"):
                continue
            parts = [p.strip() for p in line.split(',') if p.strip()]
            for i in range(0, len(parts), 8):
                if i + 7 < len(parts):
                    row = parts[i:i + 8]
                    rows.append(row)

    df = pd.DataFrame(rows, columns=["Date", "Module", "T", "P", "S", "F", "I", "KI"])
    df = df[df["Date"].str.contains(r'\d{1,2}-[A-Za-z]+-\d{4}', na=False)]
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%B-%Y")

    for col in ["T", "P", "S", "F", "I", "KI"]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    return df

def generate_graphs(df, output_dir):
    # ✅ Exclude Saturday (5) and Sunday (6)
    df = df[df["Date"].dt.weekday < 5]

    modules = df["Module"].unique()
    graph_files = []

    for module in modules:
        module_df = df[df["Module"] == module]
        module_df = module_df.groupby(["Date", "Module"]).sum().reset_index()
        module_df = module_df.sort_values("Date")

        if module_df.empty:
            continue

        plt.figure(figsize=(10, 5))
        plt.plot(module_df["Date"], module_df["T"], label="Total", linewidth=2, color='blue', marker='o')
        plt.plot(module_df["Date"], module_df["P"], label="Passed", linewidth=2.5, color='green', marker='o')
        plt.plot(module_df["Date"], module_df["F"], label="Failed", linewidth=2.5, color='red', marker='o')

        plt.title(f"Trend for Module: {module}", fontsize=14)
        plt.xlabel("Date", fontsize=12)
        plt.ylabel("Count", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend()

        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%d-%b-%Y'))
        plt.gca().xaxis.set_major_locator(mdates.DayLocator())

        plt.tight_layout()

        file_path = os.path.join(output_dir, f"{module}_trend.png")
        plt.savefig(file_path)
        graph_files.append((module, file_path))
        plt.close()

    return graph_files

def export_to_excel(csv_path, graph_files, xlsx_path):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Module Data"

    df = pd.read_csv(csv_path, header=None)
    row_data = df.values.tolist()

    num_blocks = len(row_data[0]) // 8
    headers = ["Date", "Module", "T", "P", "S", "F", "I", "KI"]
    start_col = 1

    # Write headers
    for block in range(num_blocks):
        for idx, header in enumerate(headers):
            col_letter = get_column_letter(start_col + idx)
            cell = ws_data[f"{col_letter}1"]
            cell.value = header
            cell.font = Font(bold=True)
        start_col += 9  # 8 columns + 1 gap

    # Write data
    for row_idx, row in enumerate(row_data, start=2):
        start_col = 1
        for block in range(num_blocks):
            for idx in range(8):
                col_letter = get_column_letter(start_col + idx)
                value = row[block * 8 + idx] if block * 8 + idx < len(row) else ""
                ws_data[f"{col_letter}{row_idx}"].value = value
            start_col += 9

    # Adjust column widths
    for col in ws_data.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max_length + 2

    # Add graphs in separate sheet
    ws_charts = wb.create_sheet(title="Module Graphs")
    row_pos = 1
    for module, image_path in graph_files:
        if not os.path.exists(image_path):
            continue
        img = Image(image_path)
        img.width = 800
        img.height = 400
        ws_charts.add_image(img, f"A{row_pos}")
        row_pos += 22

    wb.save(xlsx_path)

# === MAIN EXECUTION ===

csv_dir = "minio-report-tracker/csv"
output_base = "minio-report-tracker/xlxs"
os.makedirs(output_base, exist_ok=True)

for file in os.listdir(csv_dir):
    if not file.endswith(".csv"):
        continue

    alias = file.replace(".csv", "")
    csv_path = os.path.join(csv_dir, file)

    df_normalized = load_normalized_data(csv_path)

    output_dir = os.path.join(output_base, f"{alias}_images")
    os.makedirs(output_dir, exist_ok=True)

    graph_files = generate_graphs(df_normalized, output_dir)
    xlsx_path = os.path.join(output_base, f"{alias}.xlsx")

    export_to_excel(csv_path, graph_files, xlsx_path)
